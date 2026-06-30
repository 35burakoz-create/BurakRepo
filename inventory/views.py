import csv
from datetime import date
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, DetailView, FormView, ListView, TemplateView, UpdateView, View
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .forms import CustomerForm, ExcelImportUploadForm, LocationForm, OfferForm, ProductForm, ProductLocationChangeForm, ReservationForm, SaleForm, StockMovementForm
from .models import AuditLog, Customer, ImportFile, Location, Offer, Product, Reservation, StockMovement


EXCEL_COLUMN_MAP = {
    'location_part': 1,  # B
    'row_info': 2,       # C
    'material_name': 3,  # D
    'cut_direction': 4,  # E
    'surface_finish': 5, # F
    'quality': 6,        # G
    'thickness': 7,      # H
    'size': 8,           # I
    'crate_m2': 10,      # K
    'quantity_m2': 12,   # M
}


def parse_decimal_cell(value):
    if value in (None, ''):
        return Decimal('0')
    try:
        return Decimal(str(value).replace(',', '.').strip())
    except InvalidOperation:
        return None


def build_location_code(location_part, row_info):
    location_part = str(location_part or '').strip()
    row_info = str(row_info or '').strip()
    if not location_part and not row_info:
        return ''
    if '-' in location_part and row_info:
        side, section = location_part.split('-', 1)
        return f'{side.strip()}{row_info}-{section.strip()}'
    if location_part and row_info:
        return f'{location_part}{row_info}'
    return location_part or row_info


def generate_import_sku(row_number, material_name, location_code):
    base = f"IMP-{row_number}-{material_name[:12]}-{location_code[:12]}".upper()
    safe = ''.join(char if char.isalnum() else '-' for char in base).strip('-')
    candidate = safe or f'IMP-{row_number}'
    counter = 1
    unique_candidate = candidate
    while Product.objects.filter(sku=unique_candidate).exists():
        counter += 1
        unique_candidate = f'{candidate}-{counter}'
    return unique_candidate


def read_import_rows(import_file):
    path = import_file.file.path
    filename = import_file.original_name.lower()
    if filename.endswith('.xlsx'):
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        raw_rows = sheet.iter_rows(min_row=2, values_only=True)
    elif filename.endswith('.csv'):
        csv_file = open(path, newline='', encoding='utf-8-sig')
        raw_rows = csv.reader(csv_file)
        next(raw_rows, None)
    else:
        return [], [{'row_number': '-', 'messages': ['Sadece .xlsx ve .csv import desteklenir.']}]

    preview_rows = []
    errors = []
    for index, row in enumerate(raw_rows, start=2):
        values = list(row)
        if not any(values):
            continue
        def cell(name):
            position = EXCEL_COLUMN_MAP[name]
            return values[position] if len(values) > position else ''

        location_code = build_location_code(cell('location_part'), cell('row_info'))
        material_name = str(cell('material_name') or '').strip()
        quantity_m2 = parse_decimal_cell(cell('quantity_m2'))
        crate_m2 = parse_decimal_cell(cell('crate_m2'))
        row_errors = []
        if not location_code:
            row_errors.append('Konum/bölme veya sıra bilgisi yok.')
        if not material_name:
            row_errors.append('Malzeme adı yok.')
        if quantity_m2 is None or quantity_m2 <= 0:
            row_errors.append('Toplam m² geçerli ve sıfırdan büyük olmalı.')
        if crate_m2 is None:
            row_errors.append('Kasa m² geçerli olmalı.')
        parsed = {
            'row_number': index,
            'location_code': location_code,
            'material_name': material_name,
            'cut_direction': str(cell('cut_direction') or '').strip(),
            'surface_finish': str(cell('surface_finish') or '').strip(),
            'quality': str(cell('quality') or '').strip(),
            'thickness': str(cell('thickness') or '').strip(),
            'size': str(cell('size') or '').strip(),
            'crate_m2': crate_m2 if crate_m2 is not None else Decimal('0'),
            'quantity_m2': quantity_m2 if quantity_m2 is not None else Decimal('0'),
            'errors': row_errors,
        }
        preview_rows.append(parsed)
        if row_errors:
            errors.append({'row_number': index, 'messages': row_errors})
    return preview_rows, errors


def decimal_param(params, name):
    value = params.get(name, '').strip()
    if not value:
        return None
    try:
        return Decimal(value)
    except InvalidOperation:
        return None


def filtered_product_queryset(params):
    queryset = Product.objects.select_related('location')
    query = params.get('q', '').strip()
    location_id = params.get('location')

    if query:
        status_query = Q()
        for status_value, status_label in Product.StockStatus.choices:
            if query.lower() in status_label.lower():
                status_query |= Q(status=status_value)
        queryset = queryset.filter(
            Q(material_name__icontains=query)
            | Q(cut_direction__icontains=query)
            | Q(surface_finish__icontains=query)
            | Q(quality__icontains=query)
            | Q(thickness__icontains=query)
            | Q(size__icontains=query)
            | Q(location__name__icontains=query)
            | Q(status__icontains=query)
            | status_query
        )
    if location_id and location_id.isdigit():
        queryset = queryset.filter(location_id=location_id)

    filters = {
        'material_name': 'material_name__icontains',
        'thickness': 'thickness__icontains',
        'size': 'size__icontains',
        'surface_finish': 'surface_finish__icontains',
        'cut_direction': 'cut_direction__icontains',
        'quality': 'quality__icontains',
        'status': 'status',
    }
    for param, lookup in filters.items():
        value = params.get(param, '').strip()
        if value:
            queryset = queryset.filter(**{lookup: value})

    min_m2 = decimal_param(params, 'min_m2')
    max_m2 = decimal_param(params, 'max_m2')
    if min_m2 is not None:
        queryset = queryset.filter(quantity_m2__gte=min_m2)
    if max_m2 is not None:
        queryset = queryset.filter(quantity_m2__lte=max_m2)
    if params.get('only_active'):
        queryset = queryset.exclude(status=Product.StockStatus.SOLD).filter(quantity_m2__gt=0)
    products = list(queryset.order_by('location__name', 'sku'))
    if params.get('only_sellable'):
        products = [product for product in products if product.available_m2 > 0]
    return products


def filter_summary(params):
    labels = {
        'q': 'Hızlı arama',
        'material_name': 'Malzeme',
        'thickness': 'Kalınlık',
        'size': 'Ölçü',
        'surface_finish': 'Yüzey',
        'cut_direction': 'Kesim',
        'quality': 'Kalite',
        'status': 'Durum',
        'min_m2': 'Min m²',
        'max_m2': 'Max m²',
    }
    parts = []
    for key, label in labels.items():
        value = params.get(key, '').strip()
        if value:
            parts.append(f'{label}: {value}')
    location_id = params.get('location')
    if location_id and location_id.isdigit():
        location = Location.objects.filter(pk=location_id).first()
        if location:
            parts.append(f'Konum: {location.name}')
    if params.get('only_active'):
        parts.append('Sadece aktif stok')
    if params.get('only_sellable'):
        parts.append('Sadece satılabilir m²')
    return ', '.join(parts) if parts else 'Filtre yok'


class DashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'inventory.view_dashboard'
    template_name = 'inventory/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        products = Product.objects.select_related('location')
        context.update({
            'product_count': products.count(),
            'location_count': Location.objects.count(),
            'total_m2': products.aggregate(total=Sum('quantity_m2'))['total'] or 0,
            'reserved_m2': products.aggregate(total=Sum('reserved_m2'))['total'] or 0,
            'stock_value': sum(product.stock_value_usd for product in products),
            'by_location': Location.objects.annotate(product_total=Count('products')).order_by('name'),
            'recent_movements': StockMovement.objects.select_related('stock_item', 'created_by')[:8],
            'expired_reservations': Reservation.objects.select_related('stock_item').filter(status=Reservation.ReservationStatus.ACTIVE, valid_until__lt=date.today())[:8],
        })
        return context


class StockMapView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'inventory.view_product'
    template_name = 'inventory/stock_map.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        locations = {location.name: location for location in Location.objects.all()}
        products_by_location = {}
        active_products = Product.objects.select_related('location').filter(quantity_m2__gt=0).order_by('sku')
        for product in active_products:
            products_by_location.setdefault(product.location_id, []).append(product)

        rows = []
        for row_number in range(1, settings.KROKI_ROWS + 1):
            rows.append({
                'number': row_number,
                'a_cells': self._build_cells('A', row_number, settings.KROKI_A_SLOTS, locations, products_by_location),
                'b_cells': self._build_cells('B', row_number, settings.KROKI_B_SLOTS, locations, products_by_location),
            })
        context['rows'] = rows
        return context

    def _build_cells(self, side, row_number, slots, locations, products_by_location):
        cells = []
        for slot in slots:
            code = f'{side}{row_number}-{slot}'
            location = locations.get(code)
            products = products_by_location.get(location.id, []) if location else []
            cells.append({
                'code': code,
                'location': location,
                'products': products,
                'product_count': len(products),
            })
        return cells


class ProductListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'inventory.view_product'
    model = Product
    template_name = 'inventory/product_list.html'
    context_object_name = 'products'
    paginate_by = 25

    def get_queryset(self):
        return filtered_product_queryset(self.request.GET)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        location_id = self.request.GET.get('location')
        context['selected_location'] = Location.objects.filter(pk=location_id).first() if location_id and location_id.isdigit() else None
        context['locations'] = Location.objects.order_by('name')
        context['status_choices'] = Product.StockStatus.choices
        context['filter_options'] = {
            'material_names': self._distinct_values('material_name'),
            'thicknesses': self._distinct_values('thickness'),
            'sizes': self._distinct_values('size'),
            'surface_finishes': self._distinct_values('surface_finish'),
            'cut_directions': self._distinct_values('cut_direction'),
            'qualities': self._distinct_values('quality'),
        }
        return context

    def _distinct_values(self, field_name):
        return Product.objects.exclude(**{field_name: ''}).order_by(field_name).values_list(field_name, flat=True).distinct()

    def _decimal_param(self, name):
        value = self.request.GET.get(name, '').strip()
        if not value:
            return None
        try:
            return Decimal(value)
        except InvalidOperation:
            return None


class ProductDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = 'inventory.view_product'
    model = Product
    template_name = 'inventory/product_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['movements'] = self.object.movements.select_related('old_location', 'new_location', 'created_by')[:25]
        context['active_reservations'] = self.object.reservations.filter(status=Reservation.ReservationStatus.ACTIVE).select_related('created_by')
        return context


class ProductLocationChangeView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    permission_required = 'inventory.change_product'
    form_class = ProductLocationChangeForm
    template_name = 'inventory/product_location_change.html'

    def dispatch(self, request, *args, **kwargs):
        self.product = get_object_or_404(Product.objects.select_related('location'), pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        current_code = self.product.location.name
        if len(current_code) >= 4 and '-' in current_code:
            side_row, section = current_code.split('-', 1)
            initial.update({'side': side_row[:1], 'row_number': side_row[1:], 'section': section})
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = self.product
        context['current_location'] = self.product.location
        return context

    def form_valid(self, form):
        location_code = form.location_code()
        new_location, _created = Location.objects.get_or_create(name=location_code)
        old_location = self.product.location
        if old_location == new_location:
            messages.info(self.request, 'Stok zaten seçilen konumda.')
            return redirect(self.product.get_absolute_url())
        self.product.location = new_location
        self.product._current_user = self.request.user
        self.product.save(update_fields=['location', 'updated_at'])
        messages.success(self.request, f'Konum {old_location} konumundan {new_location} konumuna değiştirildi.')
        return redirect(self.product.get_absolute_url())


class ReservationCreateView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    permission_required = 'inventory.add_reservation'
    form_class = ReservationForm
    template_name = 'inventory/reservation_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.product = get_object_or_404(Product.objects.select_related('location'), pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['stock_item'] = self.product
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = self.product
        return context

    def form_valid(self, form):
        reservation = form.save(commit=False)
        reservation.stock_item = self.product
        reservation.created_by = self.request.user
        reservation._current_user = self.request.user
        reservation.save()
        messages.success(self.request, 'Rezervasyon oluşturuldu.')
        return redirect(self.product.get_absolute_url())


class ReservationCancelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'inventory.change_reservation'

    def post(self, request, *args, **kwargs):
        reservation = get_object_or_404(Reservation.objects.select_related('stock_item'), pk=kwargs['pk'])
        product = reservation.stock_item
        reservation.cancel(user=request.user)
        messages.success(request, 'Rezervasyon iptal edildi.')
        return redirect(product.get_absolute_url())


class SaleCreateView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    permission_required = 'inventory.process_sale'
    form_class = SaleForm
    template_name = 'inventory/sale_form.html'

    def dispatch(self, request, *args, **kwargs):
        product_pk = kwargs.get('pk')
        self.product = Product.objects.filter(pk=product_pk).first() if product_pk else None
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        if self.product:
            kwargs['stock_item'] = self.product
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = self.product
        return context

    def form_valid(self, form):
        product = form.cleaned_data['stock_item']
        sold_m2 = form.cleaned_data['sold_m2']
        reservation = form.cleaned_data.get('reservation')
        customer_name = form.cleaned_data['customer_name']
        note = form.cleaned_data.get('note', '')
        sale_note = f'Satış: {customer_name}'
        if note:
            sale_note = f'{sale_note} - {note}'
        if reservation:
            sale_note = f'{sale_note} (Rezervasyon #{reservation.pk})'

        with transaction.atomic():
            if reservation:
                release_m2 = min(sold_m2, reservation.reserved_m2)
                product.reserved_m2 = max(product.reserved_m2 - release_m2, 0)
                if sold_m2 >= reservation.reserved_m2:
                    reservation.reserved_m2 = 0
                    reservation.status = Reservation.ReservationStatus.CONVERTED_TO_SALE
                else:
                    reservation.reserved_m2 -= sold_m2
                reservation.save(update_fields=['reserved_m2', 'status', 'updated_at'])
            product.sold_m2 += sold_m2
            if product.available_m2 <= 0:
                product.status = Product.StockStatus.SOLD
            elif product.sold_m2 > 0:
                product.status = Product.StockStatus.PARTIALLY_SOLD
            product._current_user = self.request.user
            product._sale_note = sale_note
            product.save(update_fields=['reserved_m2', 'sold_m2', 'status', 'updated_at'])
        messages.success(self.request, 'Satış kaydedildi ve stok güncellendi.')
        return redirect(product.get_absolute_url())


class ExcelImportView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    permission_required = 'inventory.import_stock'
    form_class = ExcelImportUploadForm
    template_name = 'inventory/import_excel.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        import_id = self.request.GET.get('import_id')
        if import_id and import_id.isdigit():
            import_file = get_object_or_404(ImportFile, pk=import_id)
            preview_rows, errors = read_import_rows(import_file)
            context.update({'import_file': import_file, 'preview_rows': preview_rows[:50], 'errors': errors, 'can_import': not errors and bool(preview_rows)})
        return context

    def form_valid(self, form):
        import_file = form.save(commit=False)
        import_file.original_name = import_file.file.name
        import_file.uploaded_by = self.request.user
        import_file.save()
        return redirect(f'{self.request.path}?import_id={import_file.pk}')


class ExcelImportConfirmView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'inventory.import_stock'

    def post(self, request, *args, **kwargs):
        import_file = get_object_or_404(ImportFile, pk=kwargs['pk'])
        preview_rows, errors = read_import_rows(import_file)
        if errors:
            messages.error(request, 'Hatalı satırlar olduğu için import yapılmadı.')
            return redirect(f'{reverse_lazy("excel_import")}?import_id={import_file.pk}')
        created_count = 0
        with transaction.atomic():
            for row in preview_rows:
                location, _created = Location.objects.get_or_create(name=row['location_code'])
                product = Product(
                    sku=generate_import_sku(row['row_number'], row['material_name'], row['location_code']),
                    name=row['material_name'],
                    product_type=Product.ProductType.OTHER,
                    material_name=row['material_name'],
                    cut_direction=row['cut_direction'],
                    surface_finish=row['surface_finish'],
                    quality=row['quality'],
                    thickness=row['thickness'],
                    size=row['size'],
                    location=location,
                    crate_m2=row['crate_m2'],
                    quantity_m2=row['quantity_m2'],
                )
                product._current_user = request.user
                product.save()
                created_count += 1
            import_file.processed = True
            import_file.save(update_fields=['processed', 'updated_at'])
            if import_file.delete_after_processing:
                import_file.file.delete(save=False)
        messages.success(request, f'{created_count} stok kaydı içe aktarıldı.')
        return redirect('product_list')


class ProductCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'inventory.add_product'
    model = Product
    form_class = ProductForm
    template_name = 'inventory/form.html'
    success_url = reverse_lazy('product_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance._current_user = self.request.user
        return super().form_valid(form)



class ProductUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'inventory.change_product'
    model = Product
    form_class = ProductForm
    template_name = 'inventory/form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance._current_user = self.request.user
        return super().form_valid(form)

class CustomerListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'inventory.view_customer'
    model = Customer
    template_name = 'inventory/customer_list.html'
    context_object_name = 'customers'


class CustomerCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'inventory.add_customer'
    model = Customer
    form_class = CustomerForm
    template_name = 'inventory/form.html'
    success_url = reverse_lazy('customer_list')


class OfferListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'inventory.view_offer'
    model = Offer
    template_name = 'inventory/offer_list.html'
    context_object_name = 'offers'
    paginate_by = 25

    def get_queryset(self):
        return Offer.objects.select_related('customer', 'created_by').prefetch_related('selected_stock_items')


class OfferDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = 'inventory.view_offer'
    model = Offer
    template_name = 'inventory/offer_detail.html'

    def get_queryset(self):
        return Offer.objects.select_related('customer', 'created_by').prefetch_related('selected_stock_items__location')


class OfferCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'inventory.add_offer'
    model = Offer
    form_class = OfferForm
    template_name = 'inventory/offer_form.html'

    def get_initial(self):
        initial = super().get_initial()
        stock_item_id = self.request.GET.get('stock_item')
        if stock_item_id and stock_item_id.isdigit():
            initial['selected_stock_items'] = [stock_item_id]
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


def offer_pdf(request, pk):
    if not request.user.is_authenticated:
        return redirect('login')
    if not request.user.has_perm('inventory.view_offer'):
        return HttpResponseForbidden('Bu teklifi görüntüleme yetkiniz yok.')
    offer = get_object_or_404(Offer.objects.select_related('customer', 'created_by').prefetch_related('selected_stock_items__location'), pk=pk)
    AuditLog.objects.create(
        actor=request.user,
        action=AuditLog.Action.REPORT,
        model_name='Offer',
        object_id=str(offer.pk),
        object_repr=str(offer),
        details={'format': 'pdf', 'offer_number': offer.offer_number},
    )
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="offer_{offer.offer_number}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('NELAMAR TEKLİF TASLAĞI', styles['Title']),
        Paragraph(f'Teklif No: {offer.offer_number}', styles['Normal']),
        Paragraph(f'Oluşturma tarihi: {timezone.localtime(offer.created_at).strftime("%d.%m.%Y %H:%M")}', styles['Normal']),
        Paragraph(f'Müşteri: {offer.customer.company_name} / {offer.customer.country}', styles['Normal']),
        Paragraph(f'Para Birimi: {offer.currency} | Incoterm: {offer.incoterm or "-"}', styles['Normal']),
        Paragraph(f'Ödeme: {offer.payment_terms or "-"}', styles['Normal']),
        Paragraph(f'Geçerlilik: {offer.validity_date.strftime("%d.%m.%Y")}', styles['Normal']),
        Spacer(1, 12),
    ]
    table_data = [['Stok', 'Malzeme', 'Konum', 'Kalınlık', 'Ölçü', 'Toplam m²', 'Satılabilir m²']]
    for item in offer.selected_stock_items.all():
        table_data.append([
            item.sku,
            item.display_material_name,
            item.location.name,
            item.thickness or '-',
            item.display_size or '-',
            str(item.quantity_m2),
            str(item.available_m2),
        ])
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#212529')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph(f'Toplam m²: {offer.total_m2} | Toplam satılabilir m²: {offer.total_available_m2}', styles['Normal']))
    if offer.note:
        story.append(Paragraph(f'Not: {offer.note}', styles['Normal']))
    story.append(Spacer(1, 12))
    story.append(Paragraph('© 2026 Burak ÖZ. All rights reserved.', styles['Normal']))
    doc.build(story)
    return response


class LocationListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'inventory.view_location'
    model = Location
    template_name = 'inventory/location_list.html'
    context_object_name = 'locations'


class LocationCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'inventory.add_location'
    model = Location
    form_class = LocationForm
    template_name = 'inventory/form.html'
    success_url = reverse_lazy('location_list')


class MovementListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'inventory.view_stockmovement'
    model = StockMovement
    template_name = 'inventory/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 25

    def get_queryset(self):
        return StockMovement.objects.select_related('stock_item', 'old_location', 'new_location', 'created_by')


class MovementCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'inventory.add_stockmovement'
    model = StockMovement
    form_class = StockMovementForm
    template_name = 'inventory/form.html'
    success_url = reverse_lazy('movement_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


def stock_excel_export(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if not request.user.has_perm('inventory.export_reports'):
        return HttpResponseForbidden('Bu raporu oluşturma yetkiniz yok.')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    products = filtered_product_queryset(request.GET)
    AuditLog.objects.create(
        actor=request.user,
        action=AuditLog.Action.REPORT,
        model_name='StockExcelExport',
        details={'format': 'xlsx', 'filters': filter_summary(request.GET)},
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Stok Listesi'
    headers = [
        'Konum',
        'Malzeme',
        'Kesim yönü',
        'Yüzey işlemi',
        'Kalite',
        'Kalınlık',
        'Ölçü',
        'Kasa m²',
        'Toplam m²',
        'Rezerve m²',
        'Satılabilir m²',
        'Tahmini kasa ağırlığı',
        'Toplam tahmini ağırlık',
        'Fotoğraf albüm linki',
        'Durum',
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='212529')
    for product in products:
        sheet.append([
            product.location.name,
            product.display_material_name,
            product.cut_direction,
            product.display_surface_finish,
            product.quality,
            product.thickness,
            product.display_size,
            product.crate_m2,
            product.quantity_m2,
            product.reserved_m2,
            product.available_m2,
            product.estimated_crate_weight_kg if product.estimated_crate_weight_kg is not None else 'Katsayı tanımlı değil',
            product.estimated_total_weight_kg if product.estimated_total_weight_kg is not None else 'Katsayı tanımlı değil',
            product.photo_album_url,
            product.get_status_display(),
        ])
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="nelamar_stock_list.xlsx"'
    workbook.save(response)
    return response


def stock_pdf_report(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if not request.user.has_perm('inventory.export_reports'):
        return HttpResponseForbidden('Bu raporu oluşturma yetkiniz yok.')

    products = filtered_product_queryset(request.GET)
    include_location = bool(request.GET.get('include_location'))
    include_weights = bool(request.GET.get('include_weights'))
    include_photo_album = bool(request.GET.get('include_photo_album'))
    only_sellable = bool(request.GET.get('only_sellable'))
    filters = filter_summary(request.GET)
    total_m2 = sum((product.available_m2 if only_sellable else product.quantity_m2) for product in products)
    total_weight = None
    if include_weights:
        weights = [product.estimated_total_weight_kg for product in products if product.estimated_total_weight_kg is not None]
        total_weight = sum(weights) if weights else None

    AuditLog.objects.create(
        actor=request.user,
        action=AuditLog.Action.REPORT,
        model_name='StockReport',
        details={'format': 'pdf', 'filters': filters, 'options': dict(request.GET)},
    )
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="nelamar_stock_list.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('NELAMAR STOK LİSTESİ', styles['Title']),
        Paragraph(f'Oluşturma tarihi: {timezone.localtime().strftime("%d.%m.%Y %H:%M")}', styles['Normal']),
        Paragraph(f'Filtre bilgisi: {filters}', styles['Normal']),
        Paragraph(f'Toplam m²: {total_m2}', styles['Normal']),
        Paragraph(f'Toplam tahmini ağırlık: {total_weight if total_weight is not None else "Katsayı tanımlı olmayanlar hariç / yok"}', styles['Normal']),
        Spacer(1, 12),
    ]

    headers = ['Malzeme', 'Kesim', 'Yüzey', 'Kalite', 'Kalınlık', 'Ölçü']
    if include_location:
        headers.append('Konum')
    if not only_sellable:
        headers.append('Toplam m²')
    headers.append('Satılabilir m²')
    if include_weights:
        headers.append('Tahmini ağırlık')
    if include_photo_album:
        headers.append('Albüm')

    table_data = [[Paragraph(str(header), styles['BodyText']) for header in headers]]
    for product in products:
        row = [
            product.display_material_name,
            product.cut_direction or '-',
            product.display_surface_finish or '-',
            product.quality or '-',
            product.thickness or '-',
            product.display_size or '-',
        ]
        if include_location:
            row.append(product.location.name)
        if not only_sellable:
            row.append(str(product.quantity_m2))
        row.append(str(product.available_m2))
        if include_weights:
            row.append(f'{product.estimated_total_weight_kg} kg' if product.estimated_total_weight_kg is not None else 'Katsayı tanımlı değil')
        if include_photo_album:
            row.append(product.photo_album_url or '-')
        table_data.append([Paragraph(str(cell), styles['BodyText']) for cell in row])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#212529')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph('© 2026 Burak ÖZ. All rights reserved.', styles['Normal']))
    doc.build(story)
    return response
